import openpyxl
from ping3 import ping

# Open the Excel file and select the sheet containing the server names
workbook = openpyxl.load_workbook('servers.xlsx')
sheet = workbook['Sheet1']

# Loop through each row in the sheet, and ping each server in turn
values = [row[0] for row in sheet.iter_rows(min_row=1, values_only=True)]

for i in range(len(values)):
    server_name = values[i]
    response_time = ping(server_name)
    print(f'{server_name}')
    if response_time is None or response_time is False:
        print(f'{server_name} is down')
        sheet.cell(row=i+1, column=2).value = 'Down'
    else:
        print(f'{server_name} responded in {response_time} ms')
        sheet.cell(row=i+1, column=2).value = 'Up'
        sheet.cell(row=i+1, column=3).value = response_time


workbook.save("servers.xlsx")